import io
import pandas as pd
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.schedule import ScheduleEntry

def generate_schedule_excel(entries: List[ScheduleEntry]) -> bytes:
    """배차 결과를 이쁜 포맷의 Excel 바이너리 데이터로 변환"""

    # 1. Pandas DataFrame으로 데이터 준비
    # 종합 시간표 시트용 데이터 정렬 (출발시간 순)
    all_data = []
    for e in entries:
        all_data.append({
            "차량번호": f"{e.bus_id}호차",
            "회차": f"{e.round_no}회차",
            "구분": "기점 ➔ 종점" if e.direction == "GO" else "종점 ➔ 기점",
            "출발시간": e.departure_time,
            "도착시간": e.arrival_time,
            "대기/휴식(분)": f"{e.rest_time_after}분" if e.rest_time_after > 0 else "-"
        })
    df = pd.DataFrame(all_data)

    # 2. openpyxl Workbook 생성
    wb = Workbook()

    # --- Sheet 1: 종합 배차 시간표 ---
    ws1 = wb.active
    ws1.title = "종합 배차 시간표"
    ws1.views.sheetView[0].showGridLines = True

    # 스타일 정의
    font_family = "Malgun Gothic"
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    even_row_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 타이틀 추가
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "시내버스 일일 배차 시간표"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    # 빈 행 추가
    ws1.row_dimensions[2].height = 15

    # 헤더 작성
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws1.row_dimensions[3].height = 25

    # 데이터 작성 (출발시간 정렬을 기본으로 임시 정렬하여 기입)
    # 정렬 기준: 출발시간 문자열 기준
    df_sorted = df.sort_values(by="출발시간")
    for row_num, row_data in enumerate(df_sorted.itertuples(index=False), 4):
        ws1.row_dimensions[row_num].height = 20
        is_even = (row_num % 2 == 0)

        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 셀 채우기 (짝수행 구분)
            if is_even:
                cell.fill = even_row_fill

            # 정렬
            if col_num in [1, 2, 3]:  # 차량, 회차, 구분
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [4, 5]:   # 출발, 도착
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:                     # 휴식
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # --- Sheet 2: 차량별 순번 시각표 ---
    ws2 = wb.create_sheet(title="차량별 배차표")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "차량별 일일 운행 시간표"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40
    ws2.row_dimensions[2].height = 15

    sub_headers = ["차량", "순번", "구분", "출발시간", "도착시간", "도착후휴식", "비고"]
    for col_num, sh in enumerate(sub_headers, 1):
        cell = ws2.cell(row=3, column=col_num, value=sh)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws2.row_dimensions[3].height = 25

    # 차량별 정렬 데이터 기입 (bus_id, round_no 순)
    df_bus_sorted = df.copy()
    # 차량 번호 숫자만 추출하여 정렬하기 위해 임시 컬럼 생성
    df_bus_sorted["bus_id_num"] = df_bus_sorted["차량번호"].str.extract(r'(\d+)').astype(int)
    df_bus_sorted["round_num"] = df_bus_sorted["회차"].str.extract(r'(\d+)').astype(int)
    df_bus_sorted["dir_order"] = df_bus_sorted["구분"].apply(lambda x: 0 if "기점" in x else 1)
    df_bus_sorted = df_bus_sorted.sort_values(by=["bus_id_num", "round_num", "dir_order"])

    current_row = 4
    prev_bus = None
    bus_colors = ["FFFFFF", "F5F5F5"] # 차량 구분을 위한 배경 교차 색상
    color_idx = 0

    for row_data in df_bus_sorted.itertuples(index=False):
        ws2.row_dimensions[current_row].height = 20
        bus_id = row_data[0] # 차량번호

        # 차량이 바뀌면 배경 구분을 전환함
        if prev_bus is not None and prev_bus != bus_id:
            color_idx = (color_idx + 1) % len(bus_colors)

        current_fill = PatternFill(start_color=bus_colors[color_idx], end_color=bus_colors[color_idx], fill_type="solid")

        # 실제 표시값 채우기
        display_vals = [row_data[0], row_data[1], row_data[2], row_data[3], row_data[4], row_data[5], ""]

        for col_num, val in enumerate(display_vals, 1):
            cell = ws2.cell(row=current_row, column=col_num, value=val)
            cell.font = data_font
            cell.fill = current_fill
            cell.border = thin_border

            if col_num in [1, 2, 3, 4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        prev_bus = bus_id
        current_row += 1

    # 컬럼 너비 자동 조정 (두 시트 모두)
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # 1행 타이틀 병합 셀은 너비 계산에서 제외
            for cell in col[2:]: # 3행 헤더부터 계산
                if cell.value:
                    # 한글 인코딩 고려한 대략적인 길이 계산
                    cell_len = len(str(cell.value).encode('utf-8'))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = max(max_len // 2 + 5, 12)

    # 바이트 스트림으로 빌드 후 리턴
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.getvalue()
